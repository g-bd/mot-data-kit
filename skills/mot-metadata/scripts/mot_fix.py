#!/usr/bin/env python3
"""mot_fix - apply the mechanical fixes that a metadata audit (findings.json) points at.

Dry-run by default: prints the plan and writes fix-plan.json. `--apply` executes it after copying every
touched file to `_mot_backup_<yymmdd_hhmm>/` and writes fix-log.json. Nothing is ever deleted.

  python mot_fix.py <folder> [--findings findings.json] [--metadata file.xlsx] [--apply]
                    [--only names,encoding,cpg,metadata,columns] [--column-map map.json] [--snake-case]

Fix families (choose with --only; default = names,encoding,cpg,metadata):
  names     file names: strip invisible chars (RLM/ZWSP), spaces -> _, drop special characters, keep extension;
            shapefile sidecars renamed together; metadata Files list / Key list updated when --metadata is given
  encoding  CSV/TXT in Windows-1255 -> UTF-8 (BOM-less); original kept in backup; Data encoding updated in metadata
  cpg       write a .cpg next to every shapefile lacking one (encoding detected from the .dbf sample)
  metadata  metadata xlsx/json: key casing (version -> Version), strip invisible chars from values, Excel date cells ->
            dd/mm/yyyy text, Dataset file gets .zip, Metadata version 1.1 added, field names normalised to the file's
            spelling, duplicate Key list rows removed  -> written as <name>-fixed.xlsx (original untouched)
  columns   rename columns in CSV/XLSX headers: strip invisible chars; with --snake-case lower_snake_case for Latin
            names; with --column-map {"old":"new"} explicit renames (the only way to replace Hebrew names).
            Metadata field names follow the renames.
  zipnames  rewrite ZIP archives whose member names violate the naming rules (invisible chars, spaces, special
            characters): a new zip is written with cleaned member names, original kept in the backup.
"""
from __future__ import annotations

import argparse
import csv
import datetime as _dt
import io
import json
import re
import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from motmeta.spec import Spec, lookup_key  # noqa: E402
from motmeta.io import read_metadata, write_xlsx, write_json, to_text, as_lines  # noqa: E402
from motmeta.scan import detect_encoding, check_name, SHP_SIDECARS  # noqa: E402

INVIS = re.compile(r"[​-‏‪-‮﻿]")
DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")


def _out(m):
    try:
        print(m)
    except UnicodeEncodeError:
        print(m.encode("utf-8", "replace").decode("ascii", "replace"))


def clean_file_name(name: str) -> str:
    stem, ext = (name.rsplit(".", 1) + [""])[:2] if "." in name else (name, "")
    stem = INVIS.sub("", stem).strip()
    stem = re.sub(r"\s+", "_", stem)
    stem = re.sub(r"[*#%$&@!?<>|\"'`^~{}\[\]()+=,;:]", "", stem)
    return f"{stem}.{ext}" if ext else stem


def snake(name: str) -> str:
    n = INVIS.sub("", name).strip()
    if re.search(r"[^\x00-\x7F]", n):
        return n  # non-Latin: leave for --column-map
    n = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", "_", n)
    n = re.sub(r"[\s\-]+", "_", n)
    n = re.sub(r"[^A-Za-z0-9_]", "", n)
    return re.sub(r"_+", "_", n).lower().strip("_")


# --------------------------------------------------------------------------- planning
def plan_names(folder: Path, files: list[Path]) -> list[dict]:
    acts = []
    seen = set()
    for p in files:
        new = clean_file_name(p.name)
        if new != p.name and p.name not in seen:
            acts.append({"kind": "rename", "path": str(p.relative_to(folder)), "new": str(p.with_name(new).relative_to(folder))})
            seen.add(p.name)
    return acts


def plan_encoding(folder: Path, files: list[Path]) -> list[dict]:
    acts = []
    for p in files:
        if p.suffix.lower() not in (".csv", ".txt", ".tsv"):
            continue
        raw = p.read_bytes()[:400000]
        enc = detect_encoding(raw)
        if enc == "cp1255":
            acts.append({"kind": "reencode", "path": str(p.relative_to(folder)), "from": "cp1255", "to": "utf-8"})
    return acts


def plan_cpg(folder: Path, files: list[Path]) -> list[dict]:
    acts = []
    for p in files:
        if p.suffix.lower() == ".shp" and not p.with_suffix(".cpg").exists():
            dbf = p.with_suffix(".dbf")
            enc = "UTF-8"
            if dbf.exists():
                sample = dbf.read_bytes()[:200000]
                try:
                    sample.decode("utf-8")
                except UnicodeDecodeError:
                    enc = "1255"
            acts.append({"kind": "cpg", "path": str(p.with_suffix(".cpg").relative_to(folder)), "encoding": enc})
    return acts


def plan_columns(folder: Path, files: list[Path], snake_case: bool, colmap: dict) -> list[dict]:
    acts = []
    for p in files:
        ext = p.suffix.lower()
        if ext not in (".csv", ".txt", ".tsv", ".xlsx"):
            continue
        try:
            if ext == ".xlsx":
                import openpyxl
                wb = openpyxl.load_workbook(p, read_only=True)
                header = [to_text(c) for c in next(wb.worksheets[0].iter_rows(values_only=True), [])]
                wb.close()
            else:
                raw = p.read_bytes()[:200000]
                enc = detect_encoding(raw)
                line = raw.decode(enc, "replace").splitlines()[0]
                header = next(csv.reader([line]))
        except Exception:
            continue
        renames = {}
        for h in header:
            new = colmap.get(h) or colmap.get(INVIS.sub("", h)) or (snake(h) if snake_case else INVIS.sub("", h).strip())
            if new and new != h:
                renames[h] = new
        if renames:
            acts.append({"kind": "columns", "path": str(p.relative_to(folder)), "renames": renames})
    return acts


def plan_zipnames(folder: Path, files: list[Path]) -> list[dict]:
    import zipfile
    acts = []
    for p in files:
        if p.suffix.lower() != ".zip":
            continue
        try:
            with zipfile.ZipFile(p) as zf:
                renames = {}
                for n in zf.namelist():
                    if n.endswith("/"):
                        continue
                    parts = n.split("/")
                    new = "/".join(clean_file_name(x) for x in parts)
                    if new != n:
                        renames[n] = new
        except zipfile.BadZipFile:
            continue
        if renames:
            acts.append({"kind": "zipnames", "path": str(p.relative_to(folder)), "renames": renames})
    return acts


def plan_metadata(meta: dict, spec: Spec, file_renames: dict, col_renames: dict, enc_fixed: set) -> tuple[dict, list[dict]]:
    import copy
    m = copy.deepcopy(meta)
    acts = []
    raw_keys = m.get("_meta", {}).get("raw_keys", {})
    for k, raw in raw_keys.items():
        acts.append({"kind": "meta-key", "from": raw, "to": k})
    # invisible chars in all string values
    def strip_all(obj, path=""):
        if isinstance(obj, str):
            n = INVIS.sub("", obj)
            if n != obj:
                acts.append({"kind": "meta-invisible", "where": path})
            return n
        if isinstance(obj, list):
            return [strip_all(x, f"{path}[{i}]") for i, x in enumerate(obj)]
        if isinstance(obj, dict):
            return {k: (v if k.startswith("_") else strip_all(v, f"{path}.{k}" if path else k)) for k, v in obj.items()}
        return obj
    m = strip_all(m)
    items = spec.header_keys(True)
    for it in items:
        if it.get("format") == "date" and it["key"] in m:
            v = to_text(m[it["key"]])
            mm = re.match(r"^(\d{4})-(\d{2})-(\d{2})", v)
            if mm:
                m[it["key"]] = f"{mm.group(3)}/{mm.group(2)}/{mm.group(1)}"
                acts.append({"kind": "meta-date", "key": it["key"], "from": v, "to": m[it["key"]]})
    ds = to_text(m.get("Dataset file", ""))
    if ds and not ds.lower().endswith(".zip"):
        m["Dataset file"] = ds + ".zip"
        acts.append({"kind": "meta-zip", "from": ds, "to": m["Dataset file"]})
    if "Metadata version" not in m:
        m["Metadata version"] = spec.base["spec"]["metadata_version"]
        acts.append({"kind": "meta-add", "key": "Metadata version", "value": m["Metadata version"]})
    # file renames -> Files list / File name / Key list
    def ren(n):
        return file_renames.get(n, file_renames.get(Path(n).name, n))
    if file_renames:
        m["Files list"] = [ren(x) for x in as_lines(m.get("Files list"))]
        for fl in m.get("Files", []):
            fl["File name"] = ren(to_text(fl.get("File name")))
        m["Key list"] = [re.sub(r"^(.+?)\.([^.>]+?)\s*->\s*(.+?)\.([^.]+?)$", lambda g: f"{ren(g.group(1))}.{g.group(2)} -> {ren(g.group(3))}.{g.group(4)}", to_text(k)) for k in as_lines(m.get("Key list"))]
        acts.append({"kind": "meta-filenames", "n": len(file_renames)})
    # column renames -> field names / keys
    for fl in m.get("Files", []):
        fname = to_text(fl.get("File name"))
        rmap = col_renames.get(fname) or col_renames.get(Path(fname).name) or {}
        if fname in enc_fixed or Path(fname).name in enc_fixed:
            fl["Data encoding"] = "UTF-8"
        for fd in fl.get("File fields", []):
            n = to_text(fd.get("Name"))
            if n in rmap:
                fd["Name"] = rmap[n]
                acts.append({"kind": "meta-field", "file": fname, "from": n, "to": rmap[n]})
    # duplicate keys
    keys, seen = [], set()
    for k in as_lines(m.get("Key list")):
        kk = re.sub(r"\s+", "", to_text(k).lower())
        if kk in seen:
            acts.append({"kind": "meta-dupkey", "key": k})
            continue
        seen.add(kk)
        keys.append(to_text(k))
    if "Key list" in m:
        m["Key list"] = keys
    return m, acts


# --------------------------------------------------------------------------- apply
def apply(folder: Path, plan: dict, backup: Path) -> list[dict]:
    log = []
    backup.mkdir(parents=True, exist_ok=True)

    def bak(rel: str):
        src = folder / rel
        dst = backup / rel
        dst.parent.mkdir(parents=True, exist_ok=True)
        if src.exists() and not dst.exists():
            shutil.copy2(src, dst)

    order = {"reencode": 0, "columns": 1, "cpg": 2, "zipnames": 3, "rename": 4}   # content fixes first, renames last
    for a in sorted(plan["actions"], key=lambda x: order.get(x["kind"], 9)):
        k = a["kind"]
        try:
            if k == "reencode":
                bak(a["path"])
                p = folder / a["path"]
                text = p.read_bytes().decode("cp1255")
                p.write_bytes(text.encode("utf-8"))
            elif k == "cpg":
                (folder / a["path"]).write_text(a["encoding"], encoding="ascii")
            elif k == "columns":
                bak(a["path"])
                p = folder / a["path"]
                if p.suffix.lower() == ".xlsx":
                    import openpyxl
                    wb = openpyxl.load_workbook(p)
                    ws = wb.worksheets[0]
                    for c in ws[1]:
                        if to_text(c.value) in a["renames"]:
                            c.value = a["renames"][to_text(c.value)]
                    wb.save(p)
                else:
                    raw = p.read_bytes()
                    enc = detect_encoding(raw[:200000])
                    text = raw.decode(enc, "replace")
                    first, sep, rest = text.partition("\n")
                    hdr = next(csv.reader([first.rstrip("\r")]))
                    hdr = [a["renames"].get(h, h) for h in hdr]
                    buf = io.StringIO()
                    csv.writer(buf, lineterminator="").writerow(hdr)
                    p.write_bytes((buf.getvalue() + ("\r\n" if first.endswith("\r") else "\n") + rest).encode("utf-8" if enc in ("utf-8", "utf-8-sig") else enc))
            elif k == "zipnames":
                import zipfile
                bak(a["path"])
                p = folder / a["path"]
                tmp = p.with_suffix(".zip.tmp")
                with zipfile.ZipFile(p) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        if item.filename.endswith("/"):
                            continue
                        zout.writestr(a["renames"].get(item.filename, item.filename), zin.read(item.filename))
                p.unlink()
                tmp.rename(p)
            elif k == "rename":
                p = folder / a["path"]
                new = folder / a["new"]
                bak(a["path"])
                p.rename(new)
                if p.suffix.lower() == ".shp":  # sidecars follow
                    for sc in p.parent.iterdir():
                        if sc.with_suffix("") == p.with_suffix("") and sc.suffix.lower() in SHP_SIDECARS:
                            bak(str(sc.relative_to(folder)))
                            sc.rename(new.with_suffix(sc.suffix))
            log.append({**a, "status": "done"})
        except Exception as e:
            log.append({**a, "status": f"failed: {e}"})
    return log


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("folder")
    ap.add_argument("--findings", help="findings.json from validate (used to report which findings the plan addresses)")
    ap.add_argument("--metadata", help="metadata xlsx/json to fix alongside the files")
    ap.add_argument("--profile")
    ap.add_argument("--only", default="names,encoding,cpg,metadata,zipnames")
    ap.add_argument("--column-map", help="json {old: new} column renames")
    ap.add_argument("--snake-case", action="store_true")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--actions", help="apply only these plan numbers, e.g. \"1,3-5\" (numbers as printed in the dry run / fix-plan.json)")
    a = ap.parse_args(argv)
    folder = Path(a.folder).resolve()
    only = {x.strip() for x in a.only.split(",")}
    files = [p for p in folder.rglob("*") if p.is_file() and not any(part.startswith(("_mot_backup", ".")) for part in p.relative_to(folder).parts)
             and p.name not in ("metadata-config.json", "metadata-report.html", "findings.json", "scan.json", "fix-plan.json", "fix-log.json")]
    colmap = json.load(open(a.column_map, encoding="utf-8")) if a.column_map else {}
    actions = []
    if "names" in only:
        actions += plan_names(folder, files)
    if "encoding" in only:
        actions += plan_encoding(folder, files)
    if "cpg" in only:
        actions += plan_cpg(folder, files)
    if "columns" in only:
        actions += plan_columns(folder, files, a.snake_case, colmap)
    if "zipnames" in only:
        actions += plan_zipnames(folder, files)
    spec = Spec(a.profile)
    meta_fixed = None
    meta_acts = []
    if a.metadata and "metadata" in only:
        meta = read_metadata(Path(a.metadata), spec)
        file_renames = {Path(x["path"]).name: Path(x["new"]).name for x in actions if x["kind"] == "rename"}
        for x in actions:
            if x["kind"] == "zipnames":
                z = Path(x["path"]).name
                for old, new in x["renames"].items():
                    file_renames[f"{z}/{old}"] = f"{z}/{new}"
        col_renames = {Path(x["path"]).name: x["renames"] for x in actions if x["kind"] == "columns"}
        enc_fixed = {Path(x["path"]).name for x in actions if x["kind"] == "reencode"}
        meta_fixed, meta_acts = plan_metadata(meta, spec, file_renames, col_renames, enc_fixed)
    for i, x in enumerate(actions + meta_acts, start=1):
        x["n"] = i
    if a.actions:
        chosen: set[int] = set()
        for part in a.actions.split(","):
            part = part.strip()
            if "-" in part:
                lo, hi = part.split("-", 1)
                chosen.update(range(int(lo), int(hi) + 1))
            elif part:
                chosen.add(int(part))
        actions = [x for x in actions if x["n"] in chosen]
        meta_acts = [x for x in meta_acts if x["n"] in chosen]
        if meta_fixed is not None and not meta_acts and not any(x["kind"] in ("rename", "zipnames", "reencode", "columns") for x in actions):
            meta_fixed = None
    plan = {"folder": str(folder), "created": _dt.datetime.now().strftime("%d/%m/%Y %H:%M"), "apply": a.apply, "actions": actions, "metadata_actions": meta_acts}
    (folder / "fix-plan.json").write_text(json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8")
    _out(f"plan: {len(actions)} file actions, {len(meta_acts)} metadata actions -> fix-plan.json")
    for x in actions[:60]:
        _out(f"  {x['n']:2}. {x['kind']:9} {x.get('path')}" + (f" -> {x['new']}" if x.get("new") else "") + (f" {x['renames']}" if x.get("renames") else ""))
    for x in meta_acts[:60]:
        _out(f"  {x['n']:2}. meta      " + " ".join(f"{k}={v}" for k, v in x.items() if k not in ("kind", "n")) + f" ({x['kind']})")
    if len(actions) > 60 or len(meta_acts) > 60:
        _out(f"  ... more in fix-plan.json")
    if not a.apply:
        _out("dry run - nothing changed. Re-run with --apply to execute (backups go to _mot_backup_<stamp>/).")
        return 0
    stamp = _dt.datetime.now().strftime("%y%m%d_%H%M")
    backup = folder / f"_mot_backup_{stamp}"
    log = apply(folder, plan, backup)
    if meta_fixed is not None:
        src = Path(a.metadata)
        base = re.sub(r"[-_]?metadata$", "", src.stem) or src.stem
        out_x = src.with_name(f"{src.stem}-fixed.xlsx")
        out_j = src.with_name(f"{src.stem}-fixed.json")
        include_survey = any(k in meta_fixed for k in ("Statistical population", "Survey method"))
        meta_fixed.pop("_meta", None)
        write_xlsx(meta_fixed, out_x, spec, include_survey)
        write_json(meta_fixed, out_j)
        log.append({"kind": "metadata", "written": [str(out_x), str(out_j)], "status": "done"})
    (folder / "fix-log.json").write_text(json.dumps({"plan": plan, "log": log, "backup": str(backup)}, ensure_ascii=False, indent=2), encoding="utf-8")
    failed = [l for l in log if not str(l.get("status", "")).startswith("done")]
    _out(f"applied {len(log) - len(failed)} actions, {len(failed)} failed; backup in {backup}; log -> fix-log.json")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
