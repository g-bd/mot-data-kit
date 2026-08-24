"""Read / write the metadata document.

Canonical in-memory form (also the JSON on disk, mirrors the נוהל json example):
    {
      "Publisher": "...", "Description": ["line", "line"], "Keywords": ["a", "b"],
      ...header keys in dictionary order...,
      "Files list": ["a.csv", "b.shp"], "Key list": ["a.csv.id -> b.shp.id"],
      "Files": [ {"File name": "a.csv", "File format": "CSV", "File description": "...",
                  "File fields": [ {"Name": "id", "Type": "Integer(key)", "Description": "...",
                                    "Comments": "...", "Values": [{"value": "1", "label": "...", "comment": ""}]} ],
                  "File size": 1.2, ...}, ... ],
      "_meta": {"guideline_version": "1.3", "profile": "onboard", "raw_keys": {...}}
    }
Excel layout = the MoT example: single sheet "meta", key in column A, value(s) in column B,
block values continue on following rows with an empty A, then a "Files" section where every
file starts with "File name" and its fields follow a "File fields" header row
(name | type | description | comments | value | label | comment).
"""
from __future__ import annotations

import csv
import datetime as _dt
import html
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any, Optional

from .spec import Spec, lookup_key

FIELD_COLS = ["name", "type", "description", "comments", "value", "label", "comment"]
FIELD_KEY_BY_COL = {"name": "Name", "type": "Type", "description": "Description", "comments": "Comments"}
BLOCK_KINDS = ("block", "array")


# --------------------------------------------------------------------------- value helpers
def as_lines(v: Any) -> list[str]:
    if v is None:
        return []
    if isinstance(v, (list, tuple)):
        return [str(x) for x in v if x is not None and str(x).strip() != ""]
    s = to_text(v)
    return [s] if s != "" else []


def to_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def split_keywords(v: Any) -> list[str]:
    lines = as_lines(v)
    out: list[str] = []
    for ln in lines:
        out.extend(k.strip() for k in re.split(r"[,;\n]", ln) if k.strip())
    return out


# --------------------------------------------------------------------------- JSON
def read_json(path: Path) -> dict:
    with open(path, encoding="utf-8-sig") as f:
        txt = f.read()
    try:
        data = json.loads(txt)
    except json.JSONDecodeError:
        # tolerate trailing commas (the OB example has them)
        data = json.loads(re.sub(r",(\s*[}\]])", r"\1", txt))
    data.setdefault("_meta", {})["source_file"] = str(path)
    # normalise nested key casing
    for fl in data.get("Files", []):
        for fld in fl.get("File fields", []):
            for k in list(fld):
                if k.lower() == "name" and k != "Name":
                    fld["Name"] = fld.pop(k)
                elif k.lower() == "type" and k != "Type":
                    fld["Type"] = fld.pop(k)
                elif k.lower() == "description" and k != "Description":
                    fld["Description"] = fld.pop(k)
                elif k.lower() in ("comment", "comments") and k != "Comments":
                    fld["Comments"] = fld.pop(k)
                elif k.lower() == "values" and k != "Values":
                    fld["Values"] = fld.pop(k)
    return data


def write_json(meta: dict, path: Path) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)


# --------------------------------------------------------------------------- XLSX read
def read_xlsx(path: Path, spec: Spec) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["meta"] if "meta" in wb.sheetnames else wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header_items = spec.header_keys(include_survey=True)
    file_items = spec.file
    meta: dict[str, Any] = {"_meta": {"source_file": str(path), "sheet": ws.title, "raw_keys": {}, "unknown_keys": []}}
    files: list[dict] = []
    cur_file: Optional[dict] = None
    cur_key: Optional[str] = None
    cur_field: Optional[dict] = None
    in_fields = False
    section = "header"

    def cell(r, i):
        return r[i] if i < len(r) else None

    for r in rows:
        a, b = to_text(cell(r, 0)), cell(r, 1)
        if a == "" and all(to_text(cell(r, i)) == "" for i in range(1, max(len(r), 8))):
            continue
        if a and a.lower() == "files":
            section = "files"
            in_fields = False
            continue
        if section == "header":
            if a:
                item = lookup_key(header_items, a)
                if item is None:
                    meta["_meta"]["unknown_keys"].append(a)
                    cur_key = a
                else:
                    cur_key = item["key"]
                    if a != cur_key:
                        meta["_meta"]["raw_keys"][cur_key] = a
                val = to_text(b)
                if item and item.get("kind") in BLOCK_KINDS or item and item.get("kind") == "files":
                    meta[cur_key] = as_lines(val)
                else:
                    meta[cur_key] = val
            elif cur_key is not None:
                val = to_text(b)
                if val == "":
                    continue
                if isinstance(meta.get(cur_key), list):
                    meta[cur_key].append(val)
                else:
                    meta[cur_key] = [meta.get(cur_key, ""), val] if meta.get(cur_key, "") != "" else [val]
            continue
        # ---- files section
        if a:
            item = lookup_key(file_items, a)
            key = item["key"] if item else a
            if item and a != key:
                meta["_meta"]["raw_keys"][key] = a
            if key == "File name":
                cur_file = {"File name": to_text(b), "File fields": []}
                files.append(cur_file)
                in_fields, cur_field, cur_key = False, None, None
            elif key == "File fields":
                in_fields = True
                cur_field = None
                # header labels live in B..H; remember their order
                labels = [to_text(cell(r, i)).lower() for i in range(1, 8)]
                cur_file["_cols"] = labels if any(labels) else FIELD_COLS
            elif cur_file is not None:
                in_fields = False
                cur_key = key
                if item is None:
                    meta["_meta"]["unknown_keys"].append(f"{cur_file['File name']}:{a}")
                cur_file[key] = as_lines(to_text(b)) if (item and item.get("kind") == "block") else to_text(b)
            continue
        if cur_file is None:
            continue
        if in_fields:
            cols = cur_file.get("_cols", FIELD_COLS)
            vals = {}
            for i, lab in enumerate(cols, start=1):
                if lab:
                    vals[lab] = to_text(cell(r, i))
            name = vals.get("name", "")
            if name:
                cur_field = {"Name": name, "Type": vals.get("type", ""), "Description": vals.get("description", ""),
                             "Comments": vals.get("comments") or vals.get("comment", "") if "comments" in vals else vals.get("comment", "")}
                # when both 'comments' (field) and 'comment' (value) columns exist keep them apart
                if "comments" in vals and "comment" in vals:
                    cur_field["Comments"] = vals.get("comments", "")
                cur_file["File fields"].append(cur_field)
            if cur_field is not None and (vals.get("value", "") != "" or vals.get("label", "") != ""):
                cur_field.setdefault("Values", []).append({"value": vals.get("value", ""), "label": vals.get("label", ""),
                                                           "comment": vals.get("comment", "") if "comments" in vals else ""})
        elif cur_key and isinstance(cur_file.get(cur_key), list):
            v = to_text(b)
            if v:
                cur_file[cur_key].append(v)
    for fl in files:
        fl.pop("_cols", None)
    meta["Files"] = files
    if "Keywords" in meta:
        meta["Keywords"] = split_keywords(meta["Keywords"])
    return meta


def read_metadata(path: Path, spec: Spec) -> dict:
    path = Path(path)
    if path.suffix.lower() == ".json":
        return read_json(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return read_xlsx(path, spec)
    raise ValueError(f"unsupported metadata file type: {path.suffix}")


# --------------------------------------------------------------------------- XLSX write
def write_xlsx(meta: dict, path: Path, spec: Spec, include_survey: bool) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "meta"
    ws.sheet_view.rightToLeft = False
    bold = Font(bold=True)
    keyfill = PatternFill("solid", fgColor="E8EEF7")
    secfill = PatternFill("solid", fgColor="C9D6EA")
    hdrfill = PatternFill("solid", fgColor="F2F2F2")
    wrap = Alignment(wrap_text=True, vertical="top")
    row = 1

    def put(r, c, v, font=None, fill=None):
        cl = ws.cell(row=r, column=c, value=v)
        cl.alignment = wrap
        if font:
            cl.font = font
        if fill:
            cl.fill = fill
        return cl

    for item in spec.header_keys(include_survey):
        key = item["key"]
        if key == "Files":
            continue
        if key not in meta:
            continue
        val = meta[key]
        lines = as_lines(val) if item.get("kind") in BLOCK_KINDS or isinstance(val, list) else [to_text(val)]
        if item.get("kind") == "array" and key == "Keywords":
            lines = [", ".join(split_keywords(val))]
        put(row, 1, key, bold, keyfill)
        if not lines:
            row += 1
            continue
        for i, ln in enumerate(lines):
            put(row, 2, ln)
            row += 1
    put(row, 1, "Files", bold, secfill)
    row += 1
    file_order = [it["key"] for it in spec.file]
    for fl in meta.get("Files", []):
        for key in file_order:
            if key == "File fields":
                continue
            if key not in fl or fl[key] in ("", None, []):
                continue
            lines = as_lines(fl[key]) if isinstance(fl[key], list) else [to_text(fl[key])]
            put(row, 1, key, bold, keyfill)
            for ln in lines:
                put(row, 2, ln)
                row += 1
        # extra (profile / unknown) keys
        for key, v in fl.items():
            if key in file_order or key.startswith("_") or key == "File fields":
                continue
            put(row, 1, key, bold, keyfill)
            put(row, 2, to_text(v) if not isinstance(v, list) else "\n".join(v))
            row += 1
        put(row, 1, "File fields", bold, keyfill)
        for i, lab in enumerate(FIELD_COLS, start=2):
            put(row, i, lab, bold, hdrfill)
        row += 1
        for fld in fl.get("File fields", []):
            put(row, 2, fld.get("Name", ""))
            put(row, 3, fld.get("Type", ""))
            put(row, 4, fld.get("Description", ""))
            put(row, 5, fld.get("Comments", ""))
            values = fld.get("Values") or []
            if values:
                for j, v in enumerate(values):
                    rr = row + j
                    put(rr, 6, to_text(v.get("value", "")))
                    put(rr, 7, to_text(v.get("label", "")))
                    put(rr, 8, to_text(v.get("comment", "")))
                row += len(values)
            else:
                row += 1
        row += 1
    widths = {1: 26, 2: 48, 3: 14, 4: 48, 5: 30, 6: 12, 7: 28, 8: 24}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "B1"
    wb.save(path)


# --------------------------------------------------------------------------- CSV (key,value) write
def write_csv(meta: dict, path: Path, spec: Spec, include_survey: bool) -> None:
    """Flat key/value csv (the third format the נוהל allows)."""
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        for item in spec.header_keys(include_survey):
            key = item["key"]
            if key == "Files" or key not in meta:
                continue
            for ln in (as_lines(meta[key]) if isinstance(meta[key], list) else [to_text(meta[key])]) or [""]:
                w.writerow([key, ln])
        w.writerow(["Files", ""])
        for fl in meta.get("Files", []):
            for k, v in fl.items():
                if k == "File fields" or k.startswith("_"):
                    continue
                w.writerow([k, to_text(v) if not isinstance(v, list) else "; ".join(v)])
            w.writerow(["File fields"] + FIELD_COLS)
            for fld in fl.get("File fields", []):
                vals = fld.get("Values") or [{}]
                for j, v in enumerate(vals):
                    w.writerow(["", fld.get("Name", "") if j == 0 else "", fld.get("Type", "") if j == 0 else "",
                                fld.get("Description", "") if j == 0 else "", fld.get("Comments", "") if j == 0 else "",
                                v.get("value", ""), v.get("label", ""), v.get("comment", "")])


# --------------------------------------------------------------------------- HTML / PDF rendering
def _esc(v: Any) -> str:
    return html.escape(to_text(v))


def metadata_html(meta: dict, spec: Spec, include_survey: bool, title: Optional[str] = None) -> str:
    t = title or meta.get("Title") or "metadata"
    rows = []
    for item in spec.header_keys(include_survey):
        key = item["key"]
        if key == "Files" or key not in meta:
            continue
        val = meta[key]
        if key == "Keywords":
            body = _esc(", ".join(split_keywords(val)))
        elif isinstance(val, list):
            body = "<br>".join(_esc(x) for x in val)
        else:
            body = _esc(val)
        rows.append(f"<tr><th>{_esc(key)}<div class='he'>{_esc(item.get('he',''))}</div></th><td>{body}</td></tr>")
    files_html = []
    for fl in meta.get("Files", []):
        props = []
        for it in spec.file:
            k = it["key"]
            if k == "File fields" or k not in fl or fl[k] in ("", None, []):
                continue
            v = fl[k]
            props.append(f"<tr><th>{_esc(k)}</th><td>{'<br>'.join(_esc(x) for x in v) if isinstance(v, list) else _esc(v)}</td></tr>")
        frows = []
        for fld in fl.get("File fields", []):
            vals = fld.get("Values") or []
            vtxt = "<br>".join(f"<code>{_esc(v.get('value',''))}</code> = {_esc(v.get('label',''))}" + (f" <i>({_esc(v.get('comment'))})</i>" if v.get("comment") else "") for v in vals)
            frows.append(f"<tr><td><code>{_esc(fld.get('Name'))}</code></td><td>{_esc(fld.get('Type'))}</td><td class='he'>{_esc(fld.get('Description'))}</td><td class='he'>{_esc(fld.get('Comments'))}</td><td>{vtxt}</td></tr>")
        files_html.append(f"""
<section class='file'><h3><code>{_esc(fl.get('File name'))}</code></h3>
<table class='kv'>{''.join(props)}</table>
<table class='fields' dir='ltr'><thead><tr><th>name</th><th>type</th><th>description</th><th>comments</th><th>values</th></tr></thead><tbody>{''.join(frows)}</tbody></table>
</section>""")
    info = spec.describe()
    return f"""<!doctype html><html lang='he' dir='rtl'><head><meta charset='utf-8'><title>{_esc(t)}</title>
<style>
 @page {{ size: A4; margin: 14mm; }}
 body {{ font-family: Arial, 'Segoe UI', 'Noto Sans Hebrew', sans-serif; font-size: 11px; color:#111; margin: 0 12px; }}
 h1 {{ font-size: 18px; margin: 8px 0 2px; }} h2 {{ font-size: 14px; border-bottom: 2px solid #2b4c7e; margin: 18px 0 6px; }}
 h3 {{ font-size: 13px; margin: 14px 0 4px; color:#2b4c7e; }}
 .sub {{ color:#555; font-size: 10px; }}
 table {{ border-collapse: collapse; width: 100%; margin: 4px 0 8px; }}
 th, td {{ border: 1px solid #bbb; padding: 3px 6px; vertical-align: top; text-align: start; }}
 table.kv th {{ width: 22%; background:#eef2f8; font-weight: 600; direction: ltr; text-align: left; }}
 table.kv th .he {{ font-weight: 400; color:#666; font-size: 9.5px; direction: rtl; text-align: right; }}
 table.fields {{ direction: ltr; }} table.fields th {{ background:#f2f2f2; text-align:left; }}
 table.fields td.he {{ direction: rtl; text-align: right; }}
 code {{ font-family: Consolas, monospace; font-size: 10.5px; direction: ltr; unicode-bidi: embed; }}
 section.file {{ page-break-inside: avoid; }}
</style></head><body>
<h1>{_esc(t)}</h1>
<div class='sub'>קובץ מטא-דאטה לפי {_esc(info['guideline']['name'])} גרסה {_esc(info['guideline']['version'])}{(' · פרופיל ' + _esc(info['profile'])) if info['profile'] != 'generic' else ''}</div>
<h2>כותרת</h2><table class='kv'>{''.join(rows)}</table>
<h2>מבנה הנתונים לכל קובץ</h2>{''.join(files_html)}
</body></html>"""


def find_browsers() -> list[str]:
    cands = [
        os.environ.get("MOTMETA_BROWSER"),
        shutil.which("chrome"), shutil.which("google-chrome"), shutil.which("chromium"), shutil.which("chromium-browser"), shutil.which("msedge"),
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
    ]
    out: list[str] = []
    for c in cands:
        if c and Path(c).exists() and c not in out:
            out.append(c)
    return out


def find_browser() -> Optional[str]:
    b = find_browsers()
    return b[0] if b else None


def _wait_for_file(path: Path, timeout: float) -> bool:
    import time
    t0, last = time.time(), -1
    while time.time() - t0 < timeout:
        if path.exists():
            size = path.stat().st_size
            if size > 0 and size == last:
                return True
            last = size
        time.sleep(0.5)
    return path.exists() and path.stat().st_size > 0


def html_to_pdf(html_text: str, pdf_path: Path) -> tuple[bool, str]:
    """Render HTML to PDF with a headless Chromium (best RTL support). Tries every browser found."""
    browsers = find_browsers()
    if not browsers:
        return False, "no Chromium/Edge browser found (set MOTMETA_BROWSER) - PDF skipped, HTML kept"
    tmpdir = Path(tempfile.mkdtemp(prefix="motmeta_"))
    src = tmpdir / "doc.html"
    src.write_text(html_text, encoding="utf-8")
    pdf_path = Path(pdf_path).resolve()
    if pdf_path.exists():
        pdf_path.unlink()
    errors = []
    tmp_pdf = tmpdir / "out.pdf"          # print locally, then copy (robust on network/subst drives)
    for browser in browsers:
        cmd = [browser, "--headless=new", "--disable-gpu", "--no-first-run", "--no-default-browser-check",
               f"--user-data-dir={tmpdir / 'profile'}", "--no-pdf-header-footer",
               f"--print-to-pdf={tmp_pdf}", src.as_uri()]
        try:
            res = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
        except Exception as e:
            errors.append(f"{Path(browser).name}: {e}")
            continue
        # Chromium's launcher process often exits before the child that writes the PDF finishes
        if _wait_for_file(tmp_pdf, timeout=90):
            pdf_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(tmp_pdf, pdf_path)
            shutil.rmtree(tmpdir, ignore_errors=True)
            return True, f"pdf written with {Path(browser).name}"
        errors.append(f"{Path(browser).name}: rc={res.returncode} {res.stderr[-200:].strip()}")
    shutil.rmtree(tmpdir, ignore_errors=True)
    return False, "; ".join(errors)
