"""Folder inventory: files, sizes, dates, tabular headers + inferred types, GIS layers, ZIP contents.

Never judges the data - it only describes what is physically there so that metadata can be
generated from it or checked against it.
"""
from __future__ import annotations

import csv
import gzip
import io
import json
import os
import re
import zipfile
from collections import Counter, OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional

SAMPLE_ROWS = 5000            # rows sampled for type / cardinality inference
MAX_DISTINCT_FOR_CODED = 15   # <= this many distinct values -> candidate coded field
DATA_EXT = {".csv", ".txt", ".tsv", ".xlsx", ".xls", ".json", ".geojson", ".shp", ".gpkg", ".zip", ".gz", ".parquet", ".dbf"}
DOC_EXT = {".pdf", ".docx", ".doc", ".md", ".html", ".htm", ".rtf"}
SHP_SIDECARS = {".shx", ".dbf", ".prj", ".cpg", ".sbn", ".sbx", ".qpj", ".shp.xml", ".fix", ".qix", ".ain", ".aih", ".atx"}
METADATA_NAME_RE = re.compile(r"metadata", re.I)
KIT_OUTPUTS = {"metadata-config.json", "metadata-report.html", "findings.json", "scan.json", "package-checklist.json", "fix-plan.json", "fix-log.json"}

# A DBF declares the code page its text is written in with one byte of its header (the "language
# driver id", offset 29). This is the dBASE table - part of the file format, not of our dictionary.
# 0x57 is ESRI's "ANSI", meaning *the writing machine's* code page, which names no code page at
# all; it is deliberately absent here so that such a file falls through to DBF_FALLBACK_CODEC.
DBF_LDID_CODEPAGE = {
    0x01: "cp437", 0x02: "cp850", 0x03: "cp1252", 0x04: "mac_roman", 0x08: "cp865", 0x09: "cp437",
    0x0A: "cp850", 0x0B: "cp437", 0x0D: "cp437", 0x0E: "cp850", 0x0F: "cp437", 0x10: "cp850",
    0x11: "cp437", 0x12: "cp850", 0x13: "cp932", 0x14: "cp850", 0x15: "cp437", 0x16: "cp850",
    0x17: "cp865", 0x18: "cp437", 0x19: "cp437", 0x1A: "cp850", 0x1B: "cp437", 0x1C: "cp863",
    0x1D: "cp850", 0x1F: "cp852", 0x22: "cp852", 0x23: "cp852", 0x24: "cp860", 0x25: "cp850",
    0x26: "cp866", 0x37: "cp850", 0x40: "cp852", 0x4D: "cp936", 0x4E: "cp949", 0x4F: "cp950",
    0x50: "cp874", 0x58: "cp1252", 0x59: "cp1252", 0x64: "cp852", 0x65: "cp866", 0x66: "cp865",
    0x67: "cp861", 0x6A: "cp737", 0x6B: "cp857", 0x6C: "cp863", 0x78: "cp950", 0x79: "cp949",
    0x7A: "cp936", 0x7B: "cp932", 0x7C: "cp874", 0x86: "cp737", 0x87: "cp852", 0x88: "cp857",
    0xC8: "cp1250", 0xC9: "cp1251", 0xCA: "cp1254", 0xCB: "cp1253", 0xCC: "cp1257",
}
DBF_FALLBACK_CODEC = "cp1255"        # Windows Hebrew - what Israeli deliveries are written in
DBF_LAST_RESORT_CODEC = "cp1252"

_INT_RE = re.compile(r"^[+-]?\d+$")
_REAL_RE = re.compile(r"^[+-]?(\d+\.\d*|\.\d+|\d+)([eE][+-]?\d+)?$")
_DATE_RES = [
    (re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$"), "dd/mm/yyyy"),
    (re.compile(r"^\d{4}-\d{2}-\d{2}$"), "yyyy-mm-dd"),
    (re.compile(r"^\d{1,2}\.\d{1,2}\.\d{4}$"), "dd.mm.yyyy"),
    (re.compile(r"^\d{1,2}-\d{1,2}-\d{4}$"), "dd-mm-yyyy"),
]
_TIME_RES = [
    (re.compile(r"^\d{1,2}:\d{2}:\d{2}$"), "hh:mm:ss"),
    (re.compile(r"^\d{1,2}:\d{2}$"), "hh:mm"),
]
_DATETIME_RES = [
    (re.compile(r"^\d{4}-\d{2}-\d{2}[ T]\d{1,2}:\d{2}(:\d{2})?(\.\d+)?([+-]\d{2}:?\d{2}|Z)?$"), "yyyy-mm-dd hh:mm:ss"),
    (re.compile(r"^\d{1,2}/\d{1,2}/\d{4} \d{1,2}:\d{2}(:\d{2})?$"), "dd/mm/yyyy hh:mm:ss"),
]
NA_TOKENS = {"", "na", "n/a", "nan", "null", "none", "-", "--"}

INVISIBLE_RE = re.compile("[​-‏‪-‮⁠﻿]")
MIN_COLS_FOR_HEADER_TEST = 3      # below this a "header" is too small to judge
#: A nested zip larger than this is LISTED, not opened (KP-23: one 838 MB shapes.zip
#: was 80 % of a full-tree run). Override with MOTMETA_ZIP_MAX_MB.
NESTED_ZIP_MAX_MB = int(os.environ.get("MOTMETA_ZIP_MAX_MB", "100"))
MAX_TYPE_EXAMPLES = 5


def strip_invisible(text: str) -> str:
    return INVISIBLE_RE.sub("", str(text)).strip()


_TRAILING_Q_RE = re.compile(r"[?？]+$")
_UNDERSCORE_SPACE_RE = re.compile(r"\s*_\s*")

#: What `norm_field` had to remove, in the order it is reported. These are the ONLY
#: differences the kit treats as the same name: they are typography, not vocabulary.
#: A synonym (`boarding` for `boardings`) is deliberately NOT here - the kit does not
#: decide that one contractor's column "is" another column.
FIELD_DIRT_HE = {
    "invisible_chars": "תווים בלתי נראים (ZWSP/RLM/BOM)",
    "nbsp": "רווח קשיח (NBSP)",
    "trailing_question_mark": "סימן שאלה בסוף השם",
    "space_around_underscore": "רווח לצד קו תחתון",
    "double_space": "רווחים כפולים",
    "case": "אותיות גדולות/קטנות",
}


def norm_field(name: str) -> str:
    """A field name with the typographic dirt removed - `last _bus_stop ?` -> `last_bus_stop`."""
    n = INVISIBLE_RE.sub("", str(name)).replace(" ", " ")
    n = _TRAILING_Q_RE.sub("", n).strip()
    n = _UNDERSCORE_SPACE_RE.sub("_", n)
    return re.sub(r"\s+", " ", n).strip()


def field_key(name: str) -> str:
    """Case-folded normalised name, for matching a header against a dictionary."""
    return norm_field(name).lower()


def field_dirt(name: str) -> list[str]:
    """Which normalisations `norm_field` had to apply to *name* (empty = clean)."""
    raw = str(name)
    out: list[str] = []
    if INVISIBLE_RE.search(raw):
        out.append("invisible_chars")
    if " " in raw:
        out.append("nbsp")
    if _TRAILING_Q_RE.search(raw.strip()):
        out.append("trailing_question_mark")
    stripped = INVISIBLE_RE.sub("", raw).replace(" ", " ")
    stripped = _TRAILING_Q_RE.sub("", stripped).strip()
    if _UNDERSCORE_SPACE_RE.search(stripped) and "_" in stripped and re.search(r"\s_|_\s", stripped):
        out.append("space_around_underscore")
    if "  " in stripped:
        out.append("double_space")
    return out


def looks_like_column_name(cell: str) -> bool:
    """Could this cell be a column NAME?

    Asked the safe way round: a real header is anything that is not obviously a VALUE.
    Column names in this domain are messy - `מזרוע 1 לזרוע 2-סוג רכב 1`, `SHAPE_Leng`,
    `total boarding` - so the test rejects only what a name never is: a number, a date,
    a time, a coordinate, or a paragraph.
    """
    c = strip_invisible(cell)
    if not c or len(c) > 64:
        return False
    if _INT_RE.match(c) or _REAL_RE.match(c) or c[0].isdigit():
        return False
    for rx, _hint in (*_DATETIME_RES, *_DATE_RES, *_TIME_RES):
        if rx.match(c):
            return False
    return True


def header_is_data(header: Iterable[str]) -> bool:
    """Does this "header" look like the file's first DATA row? (KP-24)

    True when fewer than half its cells could be a column name. A file that has no
    header row must never have one of its data rows promoted to a field list - the
    fields it would invent are unanswerable by anybody.
    """
    cells = [c for c in (header or []) if strip_invisible(c)]
    if len(cells) < MIN_COLS_FOR_HEADER_TEST:
        return False
    named = sum(1 for c in cells if looks_like_column_name(c))
    return named * 2 < len(cells)


# --------------------------------------------------------------------------- helpers
def mb(n_bytes: int) -> float:
    return round(n_bytes / (1024 * 1024), 3)


def fmt_date(ts: float) -> str:
    return datetime.fromtimestamp(ts).strftime("%d/%m/%Y")


def detect_encoding(raw: bytes) -> str:
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if raw.startswith(b"\xff\xfe"):
        return "utf-16-le"
    if raw.startswith(b"\xfe\xff"):
        return "utf-16-be"
    if raw and raw[1::2].count(0) > len(raw) // 4:     # BOM-less UTF-16-LE (lots of NULs in odd positions)
        return "utf-16-le"
    if raw and raw[0::2].count(0) > len(raw) // 4:
        return "utf-16-be"
    for enc in ("utf-8", "cp1255", "latin-1"):
        try:
            raw.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "latin-1"


def infer_type(values: Iterable[str]) -> tuple[str, Optional[str]]:
    """Return (MoT type, format hint) from a sample of string cell values."""
    vals = [v.strip() for v in values if v is not None and v.strip().lower() not in NA_TOKENS]
    if not vals:
        return "Text", None
    if all(_INT_RE.match(v) for v in vals):
        return "Integer", None
    if all(_REAL_RE.match(v) for v in vals):
        return "Real", None
    for rx, hint in _DATETIME_RES:
        if all(rx.match(v) for v in vals):
            return "DateTime", hint
    for rx, hint in _DATE_RES:
        if all(rx.match(v) for v in vals):
            return "Date", hint
    for rx, hint in _TIME_RES:
        if all(rx.match(v) for v in vals):
            return "Time", hint
    return "Text", None


def _sniff_delimiter(text: str) -> str:
    head = text[:20000]
    try:
        return csv.Sniffer().sniff(head, delimiters=",;\t|").delimiter
    except csv.Error:
        counts = {d: head.count(d) for d in ",;\t|"}
        return max(counts, key=counts.get) if any(counts.values()) else ","


DISTINCT_CAP = 60             # exact distinct tracking stops once a column exceeds this


class TableProfiler:
    """Streaming column profiler: stride sample for type inference + exact (capped) distinct counts."""

    def __init__(self, header: list[str], sample_rows: int = SAMPLE_ROWS):
        self.header = header
        self.n = 0
        self.sample: list[list[str]] = []
        self.sample_rows = sample_rows
        self.distinct: list[Optional[Counter]] = [Counter() for _ in header]
        self.nulls = [0] * len(header)
        self._stride = 1

    def feed(self, row: list[str]) -> None:
        self.n += 1
        if self.n <= 200 or (self.n % self._stride == 0 and len(self.sample) < self.sample_rows):
            self.sample.append(row)
        elif len(self.sample) >= self.sample_rows:
            # thin the sample and double the stride so the sample keeps covering the whole file
            self.sample = self.sample[::2]
            self._stride *= 2
        for i, d in enumerate(self.distinct):
            v = row[i] if i < len(row) else ""
            vs = v.strip()
            if vs.lower() in NA_TOKENS:
                self.nulls[i] += 1
                continue
            if d is None:
                continue
            d[vs] += 1
            if len(d) > DISTINCT_CAP:
                self.distinct[i] = None

    def result(self) -> list[dict]:
        cols: list[dict] = []
        for i, name in enumerate(self.header):
            colvals = [r[i] if i < len(r) else "" for r in self.sample]
            mtype, hint = infer_type(colvals)
            nonnull = [v.strip() for v in colvals if v is not None and v.strip().lower() not in NA_TOKENS]
            d = self.distinct[i]
            col = {
                "name": name, "inferred_type": mtype, "format_hint": hint,
                "n_sampled": self.n, "n_null": self.nulls[i],
                "n_distinct": len(d) if d is not None else f">{DISTINCT_CAP}",
                "unique_in_sample": (d is not None and len(d) == self.n - self.nulls[i] and self.n > 1) or (d is None and len(set(nonnull)) == len(nonnull) and len(nonnull) > 1),
                "example": nonnull[0] if nonnull else None,
            }
            if mtype == "Text":
                non_numeric = [v for v in nonnull if not _REAL_RE.match(v)]
                col["text_example"] = non_numeric[0] if non_numeric else None
                # KP-8: how many of the sampled values are NOT numeric, and up to five of them,
                # so a column typed Integer that holds text is reported with evidence and a count.
                if non_numeric:
                    col["n_non_numeric_sampled"] = len(non_numeric)
                    col["text_examples"] = non_numeric[:MAX_TYPE_EXAMPLES]
            if d is not None and 0 < len(d) <= DISTINCT_CAP:
                col["distinct_values"] = sorted(d)
            if mtype in ("Date", "DateTime") and nonnull:
                col["min_value"], col["max_value"] = min(nonnull), max(nonnull)
            if d is not None and 0 < len(d) <= MAX_DISTINCT_FOR_CODED and (self.n - self.nulls[i]) > len(d):
                col["candidate_values"] = [k for k, _ in d.most_common()]
            cols.append(col)
        return cols


def profile_table(header: list[str], rows: list[list[str]]) -> list[dict]:
    tp = TableProfiler(header)
    for r in rows:
        tp.feed(r)
    return tp.result()


# --------------------------------------------------------------------------- readers
def scan_csv_bytes(raw: bytes, name: str) -> dict:
    enc = detect_encoding(raw[:200000])
    if enc.startswith("utf-16"):
        decoded = raw.decode(enc, "replace").encode("utf-8")
        return {**scan_csv_bytes(decoded, name), "encoding": enc}
    text = raw.decode(enc, errors="replace")
    delim = _sniff_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    try:
        header = next(reader)
    except StopIteration:
        return {"kind": "table", "format": "CSV", "encoding": enc, "delimiter": delim, "fields": [], "n_rows": 0, "empty": True}
    header = [h.strip().lstrip("﻿") for h in header]
    if len(header) < 2 and name.lower().endswith((".txt", ".log")):
        return {"kind": "other", "format": "TXT", "encoding": enc, "fields": [], "n_rows": 0, "note": "plain text (not a table)"}
    if header_is_data(header):
        # no header row: describe the file, never invent field names from its data (KP-24)
        n = 1
        for _ in reader:
            n += 1
        return {"kind": "table", "format": "CSV", "encoding": enc, "delimiter": delim,
                "n_rows": n, "n_cols": len(header), "fields": [], "headerless": True,
                "first_row": [strip_invisible(h) for h in header[:8]]}
    tp = TableProfiler(header)
    for row in reader:
        tp.feed(row)
    return {
        "kind": "table", "format": "CSV", "encoding": enc, "delimiter": delim,
        "n_rows": tp.n, "n_cols": len(header), "fields": tp.result(),
        "duplicate_headers": [h for h, c in Counter(header).items() if c > 1],
    }


def scan_csv(path: Path) -> dict:
    if path.suffix.lower() == ".gz":
        with gzip.open(path, "rb") as f:
            raw = f.read()
        info = scan_csv_bytes(raw, path.name)
        info["compressed"] = "gzip"
        return info
    with open(path, "rb") as f:
        return scan_csv_bytes(f.read(), path.name)


def scan_xlsx(path: Path) -> dict:
    import openpyxl  # hard dep
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if header is None:
            sheets.append({"sheet": ws.title, "fields": [], "n_rows": 0})
            continue
        header = [("" if h is None else str(h).strip()) for h in header]
        rows, n_rows = [], 0
        for row in it:
            n_rows += 1
            if len(rows) < SAMPLE_ROWS:
                rows.append([("" if v is None else (v.strftime("%d/%m/%Y") if hasattr(v, "strftime") else str(v))) for v in row])
        sheets.append({"sheet": ws.title, "n_rows": n_rows, "fields": profile_table(header, rows)})
    wb.close()
    first = sheets[0] if sheets else {"fields": [], "n_rows": 0}
    return {"kind": "table", "format": "XLSX", "sheets": sheets, "fields": first["fields"], "n_rows": first.get("n_rows", 0)}


def scan_xls(path: Path) -> dict:
    try:
        import xlrd  # optional (xlrd>=2 reads only .xls)
    except ImportError:
        return {"kind": "table", "format": "XLS", "fields": [], "error": "xlrd not installed - run setup to read legacy .xls"}
    wb = xlrd.open_workbook(str(path))
    sheets = []
    for sh in wb.sheets():
        if sh.nrows == 0:
            sheets.append({"sheet": sh.name, "fields": [], "n_rows": 0})
            continue
        header = [str(c.value).strip() for c in sh.row(0)]
        tp = TableProfiler(header)
        for i in range(1, sh.nrows):
            row = []
            for c in sh.row(i):
                if c.ctype == 3:  # date
                    try:
                        row.append(xlrd.xldate.xldate_as_datetime(c.value, wb.datemode).strftime("%d/%m/%Y"))
                    except Exception:
                        row.append(str(c.value))
                elif c.ctype == 2 and float(c.value).is_integer():
                    row.append(str(int(c.value)))
                else:
                    row.append("" if c.value is None else str(c.value))
            tp.feed(row)
        sheets.append({"sheet": sh.name, "n_rows": tp.n, "fields": tp.result()})
    first = sheets[0] if sheets else {"fields": [], "n_rows": 0}
    return {"kind": "table", "format": "XLS", "sheets": sheets, "fields": first["fields"], "n_rows": first.get("n_rows", 0)}


GPKG_GEOM = {"POINT": "Point", "MULTIPOINT": "Point", "LINESTRING": "Polyline", "MULTILINESTRING": "Polyline",
             "POLYGON": "Polygon", "MULTIPOLYGON": "Polygon"}


def scan_gpkg(path: Path) -> dict:
    """GeoPackage = SQLite: layers, geometry type, CRS, bbox and column profile via the stdlib."""
    import sqlite3
    info: dict[str, Any] = {"kind": "gis", "format": "GPKG", "layers": {}, "fields": []}
    try:
        con = sqlite3.connect(f"file:{path.as_posix()}?mode=ro", uri=True)
        con.text_factory = lambda b: b.decode("utf-8", "replace")
        cur = con.cursor()
        rows = cur.execute("SELECT table_name, data_type, min_x, min_y, max_x, max_y, srs_id FROM gpkg_contents").fetchall()
        first = True
        for tname, dtype, x0, y0, x1, y1, srs in rows:
            layer: dict[str, Any] = {"type": dtype, "srs_id": srs}
            if None not in (x0, y0, x1, y1):
                layer["bbox"] = [round(v, 6) for v in (x0, y0, x1, y1)]
            gcol = None
            if dtype == "features":
                g = cur.execute("SELECT column_name, geometry_type_name FROM gpkg_geometry_columns WHERE table_name=?", (tname,)).fetchone()
                if g:
                    gcol = g[0]
                    layer["geometry_type"] = GPKG_GEOM.get(str(g[1]).upper(), g[1])
            cols = [(r[1], r[2]) for r in cur.execute(f'PRAGMA table_info("{tname}")').fetchall()]
            header = [c for c, _ in cols if c != gcol]
            tp = TableProfiler(header)
            sel = ", ".join(f'"{c}"' for c in header) or "1"
            for r in cur.execute(f'SELECT {sel} FROM "{tname}" LIMIT {SAMPLE_ROWS}'):
                tp.feed(["" if v is None else str(v) for v in r])
            layer["n_rows"] = cur.execute(f'SELECT COUNT(*) FROM "{tname}"').fetchone()[0]
            layer["fields"] = tp.result()
            for c2, decl in cols:
                for f in layer["fields"]:
                    if f["name"] == c2 and decl:
                        d = decl.upper()
                        if "INT" in d:
                            f["inferred_type"] = "Integer"
                        elif any(x in d for x in ("REAL", "FLOA", "DOUB")):
                            f["inferred_type"] = "Real"
            info["layers"][tname] = layer
            if first and dtype == "features":
                info.update({"geometry_type": layer.get("geometry_type"), "bbox": layer.get("bbox"),
                             "n_rows": layer["n_rows"], "fields": layer["fields"],
                             "crs": {"mot_value": "EPSG:2039" if srs == 2039 else ("WGS_1984" if srs == 4326 else None), "epsg": srs}})
                first = False
        con.close()
    except Exception as e:
        info["error"] = f"gpkg read failed: {e}"
    return info


def _crs_from_prj(prj_text: str) -> dict:
    out = {"wkt": prj_text.strip()[:300]}
    try:
        from pyproj import CRS
        crs = CRS.from_wkt(prj_text)
        epsg = crs.to_epsg()
        out["epsg"] = epsg
        out["name"] = crs.name
        if epsg == 2039 or "Israel" in (crs.name or "") and "TM" in (crs.name or ""):
            out["mot_value"] = "EPSG:2039"
        elif epsg == 4326 or "WGS" in (crs.name or ""):
            out["mot_value"] = "WGS_1984"
        elif "GRS" in (crs.name or "") or epsg == 2039 - 0:
            out["mot_value"] = "GRS_1980"
        else:
            out["mot_value"] = None
    except Exception:
        t = prj_text.upper()
        out["mot_value"] = "EPSG:2039" if "ISRAEL" in t and "TM" in t else ("WGS_1984" if "WGS_1984" in t or "WGS84" in t else None)
    return out


def _cpg_codec(text: Optional[str]) -> Optional[str]:
    """Codec named by a .cpg sidecar ('UTF-8', '1255', 'cp1255', 'ISO-8859-8'...)."""
    t = (text or "").strip().strip("\x00").lower()
    if not t:
        return None
    if t.isdigit():
        t = "cp" + t
    t = t.replace(" ", "")
    try:
        import codecs
        return codecs.lookup(t).name
    except Exception:
        return None


def dbf_codec_candidates(dbf_head: bytes, cpg_text: Optional[str] = None) -> list[tuple[str, str]]:
    """Encodings to try for a DBF's text, best first, as (codec, why).

    The .cpg wins when there is one, then UTF-8 (what a modern writer produces and what pyshp
    assumes), then the code page the DBF's own header names, then Windows Hebrew. A layer whose
    attribute table cannot be decoded is a layer with NO documented fields - which is worse than
    reading it under a named assumption and saying so (KP-30).
    """
    out: list[tuple[str, str]] = []
    seen: set[str] = set()

    def _add(codec: Optional[str], why: str) -> None:
        if not codec:
            return
        key = codec.lower().replace("-", "_")
        if key in seen:
            return
        seen.add(key)
        out.append((codec, why))

    _add(_cpg_codec(cpg_text), "cpg")
    _add("utf-8", "utf-8")
    ldid = dbf_head[29] if len(dbf_head) > 29 else 0
    _add(DBF_LDID_CODEPAGE.get(ldid), f"ldid_0x{ldid:02x}")
    _add(DBF_FALLBACK_CODEC, "fallback")
    _add(DBF_LAST_RESORT_CODEC, "last_resort")
    return out


def _open_shapefile(shp_bytes: Optional[bytes] = None, shx_bytes: Optional[bytes] = None, dbf_bytes: Optional[bytes] = None,
                    path: Optional[Path] = None, cpg_text: Optional[str] = None) -> tuple[Any, dict]:
    """Open a shapefile, decoding its attribute table with the first encoding that works.

    Returns (reader, note) where note records `dbf_encoding_used`, the DBF's own `dbf_ldid`, and
    `dbf_encoding_assumed` when the encoding was neither declared in a .cpg nor plain UTF-8.
    """
    import shapefile
    head = dbf_bytes[:32] if dbf_bytes else (path.with_suffix(".dbf").read_bytes()[:32] if path and path.with_suffix(".dbf").exists() else b"")
    if path is not None and cpg_text is None:
        cpg = path.with_suffix(".cpg")
        cpg_text = cpg.read_text(errors="ignore") if cpg.exists() else None
    cands = dbf_codec_candidates(head, cpg_text)
    note: dict[str, Any] = {"dbf_ldid": head[29] if len(head) > 29 else None}
    last: Optional[Exception] = None
    for codec, why in cands:
        try:
            if path is not None:
                reader = shapefile.Reader(str(path), encoding=codec, encodingErrors="strict")
            else:
                reader = shapefile.Reader(shp=io.BytesIO(shp_bytes) if shp_bytes else None,
                                          shx=io.BytesIO(shx_bytes) if shx_bytes else None,
                                          dbf=io.BytesIO(dbf_bytes) if dbf_bytes else None,
                                          encoding=codec, encodingErrors="strict")
            for i, _rec in enumerate(reader.iterRecords()):
                if i >= 200:
                    break
            note["dbf_encoding_used"] = codec
            if why not in ("cpg", "utf-8"):
                note["dbf_encoding_assumed"] = codec
                note["dbf_encoding_reason"] = why
            return reader, note
        except UnicodeDecodeError as e:
            last = e
            continue
        except Exception as e:
            # not an encoding problem - let the caller report it as it always has
            raise e if last is None else e
    raise last if last else RuntimeError("shapefile could not be opened")


def scan_shapefile(shp_path: Path, zf: Optional[zipfile.ZipFile] = None, members: Optional[dict] = None) -> dict:
    """Scan a .shp (+ sidecars) either on disk or inside an open zip."""
    info: dict[str, Any] = {"kind": "gis", "format": "SHP", "fields": [], "sidecars": []}
    try:
        import shapefile  # pyshp
    except ImportError:
        info["error"] = "pyshp not installed - fields/bbox unavailable"
        return info
    try:
        if zf is None:
            stem = shp_path.with_suffix("")
            sidecars = [p for p in shp_path.parent.iterdir() if p.with_suffix("") == stem and p.suffix.lower() in SHP_SIDECARS]
            info["sidecars"] = [p.name for p in sidecars]
            info["size_mb_all"] = mb(shp_path.stat().st_size + sum(p.stat().st_size for p in sidecars))
            cpg = shp_path.with_suffix(".cpg")
            cpg_text = cpg.read_text(errors="ignore").strip() if cpg.exists() else None
            info["dbf_encoding_file"] = cpg_text
            reader, enc_note = _open_shapefile(path=shp_path, cpg_text=cpg_text)
            info.update(enc_note)
            prj = shp_path.with_suffix(".prj")
            prj_text = prj.read_text(errors="ignore") if prj.exists() else None
        else:
            base = shp_path.with_suffix("").as_posix()
            cpg_text = zf.read(members[".cpg"]).decode("ascii", "ignore").strip() if ".cpg" in members else None
            info["dbf_encoding_file"] = cpg_text
            reader, enc_note = _open_shapefile(shp_bytes=zf.read(members[".shp"]),
                                               shx_bytes=zf.read(members[".shx"]) if ".shx" in members else None,
                                               dbf_bytes=zf.read(members[".dbf"]) if ".dbf" in members else None,
                                               cpg_text=cpg_text)
            info.update(enc_note)
            info["sidecars"] = [Path(m).name for m in members.values()]
            prj_text = zf.read(members[".prj"]).decode("utf-8", "ignore") if ".prj" in members else None
        info["geometry_type"] = {1: "Point", 3: "Polyline", 5: "Polygon", 8: "Point", 11: "Point", 13: "Polyline", 15: "Polygon", 18: "Point", 21: "Point", 23: "Polyline", 25: "Polygon", 28: "Point", 31: "Polygon"}.get(reader.shapeType, str(reader.shapeTypeName))
        info["n_rows"] = len(reader)
        bbox = list(reader.bbox) if reader.bbox else None
        info["bbox"] = [round(v, 6) for v in bbox] if bbox else None
        info["crs"] = _crs_from_prj(prj_text) if prj_text else {"mot_value": None, "missing_prj": True}
        fields = [f for f in reader.fields if f[0] != "DeletionFlag"]
        sample = []
        for i, rec in enumerate(reader.iterRecords()):
            if i >= SAMPLE_ROWS:
                break
            sample.append(["" if v is None else str(v) for v in rec])
        header = [f[0] for f in fields]
        cols = profile_table(header, sample)
        for c, f in zip(cols, fields):
            c["dbf_type"] = {"C": "Text", "N": "Number", "F": "Real", "L": "Logical", "D": "Date"}.get(f[1], f[1])
            if f[1] == "N" and f[3] == 0:
                c["inferred_type"] = "Integer"
            elif f[1] in ("N", "F") and f[3] > 0:
                c["inferred_type"] = "Real"
        info["fields"] = cols
        info["dbf_hebrew_suspect"] = any("�" in (c.get("example") or "") for c in cols)
    except Exception as e:  # pragma: no cover
        info["error"] = f"shapefile read failed: {e}"
    return info


def scan_zip(path: Path, deep: bool = True, _depth: int = 0) -> dict:
    info: dict[str, Any] = {"kind": "archive", "format": "ZIP", "members": [], "inner": {}}
    try:
        with zipfile.ZipFile(path) as zf:
            names = [n for n in zf.namelist() if not n.endswith("/")]
            info["members"] = [{"name": n, "size_mb": mb(zf.getinfo(n).file_size)} for n in names]
            info["n_members"] = len(names)
            lower = {n.lower(): n for n in names}
            # GTFS detection
            gtfs_core = {"stops.txt", "routes.txt", "trips.txt", "stop_times.txt"}
            if gtfs_core.issubset({Path(n).name.lower() for n in names}):
                info["gtfs"] = True
            if not deep:
                return info
            # shapefiles inside
            for n in names:
                if n.lower().endswith(".shp"):
                    stem = n[:-4].lower()
                    members = {ext: lower[stem + ext] for ext in (".shp", ".shx", ".dbf", ".prj", ".cpg") if stem + ext in lower}
                    info["inner"][n] = scan_shapefile(Path(n), zf, members)
            # tabular inside (csv/txt) - limit to 20 members, 20 MB each
            count = 0
            for n in names:
                if count >= 20:
                    break
                ext = Path(n).suffix.lower()
                if info.get("gtfs") or METADATA_NAME_RE.search(Path(n).name):
                    continue
                if ext in (".csv", ".txt", ".tsv") and zf.getinfo(n).file_size < 200 * 1024 * 1024:
                    info["inner"][n] = scan_csv_bytes(zf.read(n), n)
                    count += 1
                elif ext in (".xlsx",) and zf.getinfo(n).file_size < 50 * 1024 * 1024:
                    import tempfile, os as _os
                    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
                        tf.write(zf.read(n)); tmpname = tf.name
                    try:
                        info["inner"][n] = scan_xlsx(Path(tmpname))
                    finally:
                        _os.unlink(tmpname)
                    count += 1
                elif n.lower().endswith(".csv.gz") and zf.getinfo(n).file_size < 200 * 1024 * 1024:
                    info["inner"][n] = scan_csv_bytes(gzip.decompress(zf.read(n)), n)
                    info["inner"][n]["compressed"] = "gzip"
                    count += 1
                elif ext == ".zip" and _depth < 1 and zf.getinfo(n).file_size > NESTED_ZIP_MAX_MB * 1024 * 1024:
                    # too big to open: list it and say so, do not walk it (KP-23)
                    info.setdefault("nested_zips", []).append(
                        {"name": n, "size_mb": mb(zf.getinfo(n).file_size), "listed_only": True,
                         "reason": f"nested zip larger than {NESTED_ZIP_MAX_MB} MB - listed, not opened"})
                    info["nested_zip_cap_mb"] = NESTED_ZIP_MAX_MB
                elif ext == ".zip" and _depth < 1:
                    import tempfile as _tf
                    import os as _os
                    with _tf.NamedTemporaryFile(suffix=".zip", delete=False) as tf:
                        tf.write(zf.read(n))
                        tmpz = tf.name
                    try:
                        sub = scan_zip(Path(tmpz), deep=deep, _depth=_depth + 1)
                    finally:
                        _os.unlink(tmpz)
                    for m, mi in (sub.get("inner") or {}).items():
                        info["inner"][f"{n}/{m}"] = mi     # flatten: "inner.zip/member"
                    info.setdefault("nested_zips", []).append({"name": n, "n_members": sub.get("n_members", 0), "gtfs": sub.get("gtfs", False)})
                    count += 1
    except zipfile.BadZipFile:
        info["error"] = "not a valid zip"
    return info


def scan_json(path: Path) -> dict:
    try:
        with open(path, encoding="utf-8-sig") as f:
            data = json.load(f)
    except Exception as e:
        return {"kind": "other", "format": "JSON", "error": str(e)}
    info: dict[str, Any] = {"kind": "other", "format": "JSON"}
    if isinstance(data, dict) and data.get("type") == "FeatureCollection":
        feats = data.get("features", [])
        props = OrderedDict()
        for ft in feats[:SAMPLE_ROWS]:
            for k, v in (ft.get("properties") or {}).items():
                props.setdefault(k, []).append("" if v is None else str(v))
        info.update({"kind": "gis", "format": "GeoJSON", "n_rows": len(feats),
                     "geometry_type": (feats[0]["geometry"]["type"] if feats and feats[0].get("geometry") else None),
                     "fields": profile_table(list(props), [list(x) for x in zip(*props.values())] if props else []),
                     "crs": {"mot_value": "WGS_1984"}})
    elif isinstance(data, list) and data and isinstance(data[0], dict):
        keys = list(OrderedDict((k, None) for d in data[:SAMPLE_ROWS] for k in d))
        rows = [["" if d.get(k) is None else str(d.get(k)) for k in keys] for d in data[:SAMPLE_ROWS]]
        info.update({"kind": "table", "n_rows": len(data), "fields": profile_table(keys, rows)})
    elif isinstance(data, dict):
        info["top_keys"] = list(data)[:50]
        if METADATA_NAME_RE.search(path.name):
            info["looks_like_metadata"] = True
    return info


# --------------------------------------------------------------------------- main entry
def scan_folder(folder: str | Path, recursive: bool = True, deep_zip: bool = True, exclude: Iterable[str] = (), docs: bool = True) -> dict:
    folder = Path(folder).resolve()
    if not folder.is_dir():
        raise NotADirectoryError(str(folder))
    exclude = set(exclude) | KIT_OUTPUTS
    entries: list[dict] = []
    paths = sorted(folder.rglob("*") if recursive else folder.iterdir(), key=lambda p: p.name.lower())
    seen_shp_stems: set[str] = set()
    for p in paths:
        if p.is_dir() or p.name in exclude or p.name.startswith(("~$", ".")):
            continue
        rel = p.relative_to(folder).as_posix()
        parts = Path(rel).parts
        if rel in exclude or any(part in exclude or part.startswith((".", "_mot_backup")) for part in parts[:-1])                 or any("/".join(parts[:i]) in exclude for i in range(1, len(parts))):
            continue
        ext = p.suffix.lower()
        if p.name.lower().endswith(".csv.gz"):
            ext = ".gz"
        st = p.stat()
        entry: dict[str, Any] = {"name": rel, "ext": ext, "size_mb": mb(st.st_size), "modified": fmt_date(st.st_mtime)}
        # shapefile sidecars are folded into the .shp entry
        if ext in SHP_SIDECARS or p.name.lower().endswith(".shp.xml"):
            stem = p.with_suffix("").as_posix() if not p.name.lower().endswith(".shp.xml") else p.name[:-8]
            entry.update({"role": "sidecar", "of": Path(stem).with_suffix(".shp").name})
            entries.append(entry)
            continue
        if METADATA_NAME_RE.search(p.name) and ext in (".xlsx", ".json", ".csv", ".pdf"):
            entry["role"] = "metadata"
        elif ext in DOC_EXT:
            entry["role"] = "document"
        elif ext in DATA_EXT:
            entry["role"] = "data"
        else:
            entry["role"] = "other"
        try:
            if entry["role"] == "data" or (entry["role"] == "metadata" and ext != ".pdf"):
                if ext in (".csv", ".txt", ".tsv", ".gz"):
                    entry.update(scan_csv(p))
                elif ext == ".xlsx":
                    entry.update(scan_xlsx(p))
                elif ext == ".shp":
                    entry.update(scan_shapefile(p))
                    seen_shp_stems.add(p.with_suffix("").as_posix())
                elif ext == ".zip":
                    entry.update(scan_zip(p, deep=deep_zip))
                elif ext in (".json", ".geojson"):
                    entry.update(scan_json(p))
                elif ext == ".gpkg":
                    entry.update(scan_gpkg(p))
                elif ext == ".xls":
                    entry.update(scan_xls(p))
        except Exception as e:  # keep scanning
            entry["error"] = f"{type(e).__name__}: {e}"
        if entry.get("kind") == "other" and entry.get("format") == "TXT":
            entry["role"] = "document"
        entries.append(entry)
    data_files = [e for e in entries if e["role"] == "data"]
    subdirs = [] if recursive else sorted(p.name for p in folder.iterdir() if p.is_dir() and not p.name.startswith("."))
    result = {
        "subfolders_not_scanned": subdirs,
        "folder": str(folder),
        "scanned_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "recursive": recursive,
        "n_files": len(entries),
        "n_data_files": len(data_files),
        "total_size_mb": round(sum(e["size_mb"] for e in entries), 2),
        "metadata_files": [e["name"] for e in entries if e["role"] == "metadata"],
        "documents": [e["name"] for e in entries if e["role"] == "document"],
        "files": entries,
    }
    if docs:
        try:
            from .docs import harvest
            harvest(folder, result)
        except Exception as e:  # never fail a scan because of documentation parsing
            result["doc_harvest"] = {"error": f"{type(e).__name__}: {e}"}
    return result


def read_column_values(folder: Path, rel: str, column: str, limit: int = 200000) -> Optional[set]:
    """Distinct values of one column of one file. None when unreadable."""
    c = read_column_counts(folder, rel, column, limit)
    return None if c is None else set(c)


def read_column_counts(folder: Path, rel: str, column: str, limit: int = 200000) -> Optional[Counter]:
    """How many rows carry each value of one column (top level, or a member of a zip).

    None when unreadable. `limit` caps the DISTINCT values, exactly as it always did for
    `read_column_values` - a join on a multi-million-row file must see the same key set it saw
    before counting was added to this function.
    """
    rel = rel.replace("\\", "/")
    try:
        if "/" in rel and not (folder / rel).exists():
            zip_part, _, member = rel.partition("/")
            zp = folder / zip_part
            if not zp.exists() or zp.suffix.lower() != ".zip":
                return None
            raw = _zip_member_bytes(zp.read_bytes(), member)
            if raw is None:
                return None
            if member.lower().endswith(".gz"):
                raw = gzip.decompress(raw)
            return _values_from_csv_bytes(raw, column, limit)
        p = folder / rel
        if not p.exists():
            return None
        ext = p.suffix.lower()
        if ext in (".csv", ".txt", ".tsv"):
            return _values_from_csv_bytes(p.read_bytes(), column, limit)
        if ext == ".gz":
            with gzip.open(p, "rb") as f:
                return _values_from_csv_bytes(f.read(), column, limit)
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            it = ws.iter_rows(values_only=True)
            header = [("" if h is None else str(h).strip()) for h in next(it, [])]
            if column not in header:
                wb.close()
                return None
            i = header.index(column)
            out: Counter = Counter()
            for row in it:
                if len(out) >= limit:
                    break
                v = row[i] if i < len(row) else None
                if v is not None and str(v).strip() != "":
                    out[str(v).strip()] += 1
            wb.close()
            return out
        if ext == ".shp":
            r, _note = _open_shapefile(path=p)
            names = [f[0] for f in r.fields if f[0] != "DeletionFlag"]
            if column not in names:
                return None
            i = names.index(column)
            out = Counter()
            for rec in r.iterRecords():
                if len(out) >= limit:
                    break
                v = rec[i]
                if v is not None and str(v).strip() != "":
                    out[str(v).strip()] += 1
            return out
        if ext == ".zip":
            with zipfile.ZipFile(p) as zf:
                names = zf.namelist()
                shp = next((n for n in names if n.lower().endswith(".shp")), None)
                if shp:
                    vals = _values_from_shp_bytes(zf, shp, column, limit)
                    if vals is not None:
                        return vals
                for n in names:
                    if n.lower().endswith((".csv", ".txt", ".tsv")):
                        vals = _values_from_csv_bytes(zf.read(n), column, limit)
                        if vals:
                            return vals
            return None
    except Exception:
        return None
    return None


def _zip_member_bytes(zip_bytes: bytes, member: str) -> Optional[bytes]:
    """Read a member from zip bytes; `member` may point through nested zips ("inner.zip/file.csv")."""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = {n.lower(): n for n in zf.namelist()}
        real = names.get(member.lower())
        if real is not None:
            return zf.read(real)
        parts = member.split("/")
        for i in range(1, len(parts)):
            head, rest = "/".join(parts[:i]), "/".join(parts[i:])
            cand = names.get(head.lower())
            if cand is not None and head.lower().endswith(".zip"):
                return _zip_member_bytes(zf.read(cand), rest)
    return None


def _values_from_shp_bytes(zf: zipfile.ZipFile, shp_name: str, column: str, limit: int) -> Optional[Counter]:
    try:
        import shapefile  # noqa: F401
    except ImportError:
        return None
    stem = shp_name[:-4].lower()
    lower = {n.lower(): n for n in zf.namelist()}
    try:
        rd, _note = _open_shapefile(
            shp_bytes=zf.read(lower[stem + ".shp"]),
            shx_bytes=zf.read(lower[stem + ".shx"]) if stem + ".shx" in lower else None,
            dbf_bytes=zf.read(lower[stem + ".dbf"]) if stem + ".dbf" in lower else None,
            cpg_text=zf.read(lower[stem + ".cpg"]).decode("ascii", "ignore") if stem + ".cpg" in lower else None,
        )
        names = [f[0] for f in rd.fields if f[0] != "DeletionFlag"]
        if column not in names:
            return None
        i = names.index(column)
        out: Counter = Counter()
        for rec in rd.iterRecords():
            if len(out) >= limit:
                break
            v = rec[i]
            if v is not None and str(v).strip() != "":
                out[str(v).strip()] += 1
        return out
    except Exception:
        return None


def _values_from_csv_bytes(raw: bytes, column: str, limit: int) -> Optional[Counter]:
    enc = detect_encoding(raw[:200000])
    text = raw.decode(enc, "replace")
    delim = _sniff_delimiter(text)
    rd = csv.reader(io.StringIO(text), delimiter=delim)
    header = next(rd, None)
    if not header:
        return None
    header = [h.strip().lstrip("\ufeff") for h in header]
    norm = {re.sub(r"[\u200b-\u200f\u202a-\u202e\ufeff]", "", h).strip().lower(): i for i, h in enumerate(header)}
    i = norm.get(re.sub(r"[\u200b-\u200f\u202a-\u202e\ufeff]", "", column).strip().lower())
    if i is None:
        return None
    out: Counter = Counter()
    for row in rd:
        if len(out) >= limit:
            break
        if i < len(row):
            v = row[i].strip()
            if v and v.lower() not in NA_TOKENS:
                out[v] += 1
    return out


def check_name(name: str) -> list[str]:
    """Return naming-rule violations (נוהל 5.7) for a file or field name."""
    problems = []
    stem = name
    if re.search(r"[֐-׿]", stem):
        problems.append("hebrew_letters")
    elif re.search(r"[^\x00-\x7F]", stem):
        problems.append("non_latin")
    if " " in stem:
        problems.append("spaces")
    if re.search(r"[*#%$&@!?<>|\"'`^~{}\[\]()+=,;:]", stem):
        problems.append("special_chars")
    if "​" in stem or "‏" in stem or "‎" in stem:
        problems.append("invisible_chars")
    return problems


def name_style(names: list[str]) -> dict:
    """Detect compound-word styles used by a list of field names."""
    styles = Counter()
    for n in names:
        core = re.sub(r"[^A-Za-z0-9_]", "", n)
        if not core:
            continue
        if "_" in core:
            styles["snake_case" if core == core.lower() else "Snake_Case"] += 1
        elif re.match(r"^[a-z]+[A-Z]", core):
            styles["camelCase"] += 1
        elif re.match(r"^[A-Z][a-z]+[A-Z]", core):
            styles["CamelCase"] += 1
        elif core.isupper() and len(core) > 3:
            styles["UPPER"] += 1
        else:
            styles["single_word"] += 1
    return dict(styles)
